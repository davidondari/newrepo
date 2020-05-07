import openpyxl
import pprint

wb = openpyxl.load_workbook("censuspopdata.xlsx")
sheet = wb.active
census = {}

for row in range(2, int(sheet.max_row) + 1):
    state = sheet.cell(row, 2).value
    county = sheet.cell(row, 3).value
    pop = sheet.cell(row, 4).value

    # make sure that the key for the state exists.
    census.setdefault(state, {})
    # make sure the key for the count exists
    census[state].setdefault(county, {"track": 0, "pop": 0})
    # making sure the population and track exists
    # increase the tracks
    census[state][county]["track"] += 1
    # sum up the population
    census[state][county]["pop"] += int(pop)
    # open text file and input the dictionary data

textfile = open("census.py", "w")
textfile.write("all_data =" + pprint.pformat(census))
textfile.close()
print("done...")
